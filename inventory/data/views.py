from django.views import View
from django.views.generic import (
    ListView, DetailView, CreateView, 
    UpdateView, DeleteView
)
from django.core.paginator import Paginator
from django.urls import reverse_lazy
from django.utils.translation import gettext_lazy as _
from .models import BaseInfo, InventoryItem
from .forms import BaseInfoForm, InventoryItemForm
from .filters import InventoryItemFilter
from django_filters.views import FilterView
from django.http import FileResponse
import os
from django.conf import settings
from .create_exel import *
from django.shortcuts import render, redirect
from django.contrib import messages
from .import_utils import import_inventory_from_excel
import os
from django.conf import settings
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import io
import pandas as pd
from django.core.files.base import ContentFile
from django.urls import reverse


def download_template(request):
    # Генерируем файл
    template_path = generate_excel_template()
    
    # Отправляем файл пользователю
    response = FileResponse(open(template_path, 'rb'))
    response['Content-Disposition'] = f'attachment; filename="inventory_template.xlsx"'
    
    # Удаляем временный файл после отправки
    def cleanup():
        os.remove(template_path)
    
    response.closed = cleanup
    return response

class BaseInfoListView(ListView):
    model = BaseInfo
    template_name = 'baseinfo_list.html'
    context_object_name = 'bases'
    paginate_by = 10
    ordering = ['base_name']
    

class BaseInfoDetailView(DetailView):
    model = BaseInfo
    template_name = 'baseinfo_detail.html'
    context_object_name = 'base'
    
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        inventory_items = self.object.inventory_items.all()
        paginator = Paginator(inventory_items, 10)  # 10 объектов на страницу
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context['page_obj'] = page_obj
        context['is_paginated'] = paginator.num_pages > 1
        return context

class BaseInfoCreateView(CreateView):
    model = BaseInfo
    form_class = BaseInfoForm
    template_name = 'baseinfo_form.html'
    success_url = reverse_lazy('baseinfo-list')

class BaseInfoUpdateView(UpdateView):
    model = BaseInfo
    form_class = BaseInfoForm
    template_name = 'baseinfo_form.html'
    success_url = reverse_lazy('baseinfo-list')

class BaseInfoDeleteView(DeleteView):
    model = BaseInfo
    template_name = 'baseinfo_confirm_delete.html'
    success_url = reverse_lazy('baseinfo-list')

class InventoryItemListView(FilterView):
    model = InventoryItem
    template_name = 'inventoryitem_list.html'
    context_object_name = 'items'
    filterset_class = InventoryItemFilter
    paginate_by = 20
    
    def get_paginate_by(self, queryset):
        # Получаем значение per_page из GET-параметров
        return self.request.GET.get('per_page', self.paginate_by)
    
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Добавляем фильтр в контекст для формы фильтрации
        context['filter'] = context['filter']
        return context



class InventoryItemDetailView(DetailView):
    model = InventoryItem
    template_name = 'inventoryitem_detail.html'
    context_object_name = 'item'

class InventoryItemCreateView(CreateView):
    model = InventoryItem
    form_class = InventoryItemForm
    template_name = 'inventoryitem_form.html'
    
    def get_success_url(self):
        return reverse_lazy('inventoryitem-detail', kwargs={'pk': self.object.pk})

class InventoryItemUpdateView(UpdateView):
    model = InventoryItem
    form_class = InventoryItemForm
    template_name = 'inventoryitem_form.html'
    
    def get_success_url(self):
        return reverse_lazy('inventoryitem-detail', kwargs={'pk': self.object.pk})

class InventoryItemDeleteView(DeleteView):
    model = InventoryItem
    template_name = 'inventoryitem_confirm_delete.html'
    
    
    def get_success_url(self):
        base = self.object.base
        return reverse_lazy('baseinfo-detail', kwargs={'pk': base.pk})

    
def import_inventory(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "Файл не был загружен")
            return redirect('import-inventory')
        
        try:
            # Сохраняем временный файл
            temp_path = os.path.join(settings.MEDIA_ROOT, 'temp', excel_file.name)
            os.makedirs(os.path.dirname(temp_path), exist_ok=True)
            
            with open(temp_path, 'wb+') as destination:
                for chunk in excel_file.chunks():
                    destination.write(chunk)
            
            # Импортируем данные
            result = import_inventory_from_excel(temp_path)
            
            # Удаляем временный файл
            os.remove(temp_path)
            
            # Формируем сообщение о результате
            success_msg = (
                f"Импорт завершен. Добавлено: {result['created']}, "
                f"Обновлено: {result['updated']}, "
                f"Всего обработано: {result['total']}"
            )
            messages.success(request, success_msg)
            
            if result['errors']:
                for error in result['errors']:
                    messages.warning(request, error)
            
            return redirect('inventoryitem-list')
            
        except Exception as e:
            messages.error(request, f"Ошибка при обработке файла: {str(e)}")
            return redirect('import-inventory')
    
    return render(request, 'import_inventory.html')



class ExportBaseInfoExcel(View):
    def get(self, request, pk):
        base = BaseInfo.objects.get(pk=pk)
        inventory_items = base.inventory_items.all()
        
        # Создаем Excel-файл
        wb = Workbook()
        ws = wb.active
        ws.title = f"Инвентарь {base.base_name}"
        
        # Заголовки
        headers = [
            "Инв. номер", "Наименование", "Кабинет", 
            "Состояние", "Дата ввода", "QR-код"
        ]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header).font = Font(bold=True)
        
        # Данные
        for row_num, item in enumerate(inventory_items, 2):
            ws.cell(row=row_num, column=1, value=item.inventory_number)
            ws.cell(row=row_num, column=2, value=item.objects_name)
            ws.cell(row=row_num, column=3, value=item.office)
            ws.cell(row=row_num, column=4, value=item.get_state_display())
            ws.cell(row=row_num, column=5, value=item.start_date.strftime("%d.%m.%Y"))
            
            # Добавляем QR-код если есть
            if hasattr(item, 'qr_code') and item.qr_code.qr:
                qr_img = item.qr_code.qr
                img = Image(io.BytesIO(qr_img.read()))
                img.width = 100
                img.height = 100
                ws.add_image(img, f'F{row_num}')
        
        # Настраиваем ширину столбцов
        column_widths = [15, 40, 10, 15, 15, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Сохраняем в HttpResponse
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="inventory_{base.base_name}.xlsx"'
        wb.save(response)
        
        return response
    
    