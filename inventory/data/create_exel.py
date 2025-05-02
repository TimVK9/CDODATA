import openpyxl
from openpyxl.styles import Font
from datetime import datetime

def generate_excel_template():
    # Создаем новую книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Инвентаризация"
    
    # Заголовки столбцов
    headers = [
        "Наименование объекта",
        "Инвентарный номер* (12 цифр)",
        "Дата ввода",
        "Состояние",
        "Счет актива",
        "Ответственный за содержание",
        "Подразделение",
        "Кабинет",
        "Описание"
    ]
    
    # Записываем заголовки
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    # Примеры данных для подсказок
    example_data = [
        "Проектор BENQ MP",
        "000000000364",
        datetime.now().strftime("%Y-%m-%d"),
        "Введено в эксплуатацию",
        "101.24, Машины и оборудование – особо ценное движимое имущество учреждения",
        "Иванов Иван Иванович",
        "Центр информационных технологий",
        "205",
        "Проектор Full HD, 3500 люмен, серийный номер XYZ123"
    ]
    
    # Записываем пример данных
    for col_num, value in enumerate(example_data, 1):
        ws.cell(row=2, column=col_num, value=value)
    
    # Добавляем пояснения
    notes = [
        "*Обязательные поля",
        "Инвентарный номер должен содержать ровно 12 цифр (дополняется нулями слева)",
        "Дата в формате ГГГГ-ММ-ДД",
        "Состояние: 'Введено в эксплуатацию' или 'Снято с учёта'",
        "Счет актива должен соответствовать существующим значениям",
        "Подразделение должно существовать в системе или будет создано автоматически"
    ]
    
    for row_num, note in enumerate(notes, 4):
        ws.cell(row=row_num, column=1, value=note)
    
    # Настраиваем ширину столбцов
    column_widths = [35, 25, 15, 25, 60, 30, 30, 10, 50]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Сохраняем файл
    file_path = "inventory_template.xlsx"
    wb.save(file_path)
    return file_path